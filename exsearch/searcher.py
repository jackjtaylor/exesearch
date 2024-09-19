"""
This searches through worksheets by Excel to find search terms, by using a Query.
"""

from query import Query
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from pathlib import Path
from io import BytesIO
from msoffcrypto import OfficeFile
from collections import defaultdict


class Searcher:
    def __init__(self) -> None:
        """
        This initialises the class.
        """
        pass

    def run_query(self, query: Query):
        """
        This function searches through each file in a path and if valid, searches that workbook for
        the query term.

        :param search: The query to perform.
        :type search: Query
        """
        for file in query.path.rglob("*.xlsm" and "*.xlsx"):
            if file.is_file and "$" not in file.name:
                file_path = Path(query.path, file)
                self.search_for_term_in_workbook(file_path, query)

    def prepare_workbook(self, book_path: Path) -> BytesIO:
        """
        This function prepares a workbook to be edited, by unlocking any encrypted files.

        :param book_path: The path of the workbook to prepare
        :type book_path: Path
        :return: The bytes stream of the decrypted workbook
        :rtype: BytesIO
        """
        with open(book_path, "rb") as workbook:
            office_file = OfficeFile(workbook)
            workbook.seek(0)  # This resets the reader for the next workbook.

            unencrypted_type = "plain"

            if office_file.is_encrypted() and office_file.type != unencrypted_type:
                return self.decrypt_workbook(office_file)

            else:
                return BytesIO(workbook.read())  # If the workbook is unencrypted, create a stream.

    def decrypt_workbook(self, office_file: OfficeFile) -> BytesIO:  # type: ignore
        """
        This is a function that decrypts a workbook using a password provided by the user. If the user
        doesn't enter the correct password after three attempts, the workbook is returned empty.

        :param office_file: The file to decrypt.
        :type office_file: OfficeFile
        :return: The decrypted file's byte stream.
        :rtype: OfficeFile
        """
        decrypted_workbook = (
            BytesIO()
        )  # This creates an in-memory BytesIO object to write the file to.

        for _ in range(3):  # This gives the user three tries to decrypt the workbook.
            key = input("Please enter the password to decrypt a file.")

            try:
                office_file.load_key(password=key)
                office_file.decrypt(decrypted_workbook)
            except:
                continue

        if not office_file.is_encrypted():  # If the stream has content and was decrypted.
            return decrypted_workbook

        else:
            raise ImportError("The file could not be decrypted.")

    def search_for_term_in_workbook(self, file_path: Path, query: Query):
        """
        This function finds terms inside a book, in all sheets.

        :param book: The path to the book
        :type book: Path
        :return: The count of how many times a term was found
        :rtype: int
        """
        workbook_byte_stream = self.prepare_workbook(file_path)

        workbook: Workbook = load_workbook(
            filename=workbook_byte_stream, data_only=True, read_only=True
        )
        self.print_workbook(file_path)  # This prints the name and path of the workbook

        for sheet in workbook:
            self.search_for_term_in_sheet(sheet, query)

        self.print_found_cells(query.matches)

    def search_for_term_in_sheet(self, sheet: Worksheet, query: Query):
        """
        This function searches for a term inside a sheet, within a workbook.

        :param count: _description_
        :type count: _type_
        :param found_cells: _description_
        :type found_cells: _type_
        :param sheet: _description_
        :type sheet: _type_
        """
        for column in sheet.iter_rows():
            for cell in column:
                if query.is_exclusive:  # If the search is exclusive, match the exact term.
                    if query.term == str(cell.value):
                        self.add_to_found(sheet, cell, query)
                else:
                    if (
                        query.is_case_sensitive
                    ):  # If the search is just case sensitive, match words.
                        if query.term in str(cell.value):
                            self.add_to_found(sheet, cell, query)
                    else:  # If the search only needs the same characters, match anything.
                        if query.term.lower().strip() in str(cell.value).lower().strip():
                            self.add_to_found(sheet, cell, query)

    def add_to_found(self, sheet: Worksheet, cell: Cell, query: Query):
        """
        This function adds a cell to the found cells in a search.

        :param sheet: The worksheet the cell was found in.
        :type sheet: Worksheet
        :param cell: The cell that was found to match.
        :type cell: Cell
        :param search: The search to add this result to.
        :type search: Query
        """
        query.matches[sheet.title] += f"{cell.column_letter}{cell.row}, "

    def print_workbook(self, book_path: Path) -> None:
        """
        This function prints the workbook name and path.

        :param book_path: The path to the workbook
        :type book_path: Path
        """
        print(f"\n_______________ Book: {book_path.name} _______________\n")
        print(f"Path: {book_path.absolute()}\n")

    def print_found_cells(self, sheet_cells: defaultdict[str, str]) -> None:
        """
        This function prints the cells found, grouped by sheet.

        :param sheet_cells: The cells found, by sheet
        :type sheet_cells: defaultdict[str, str]
        """
        print("Cells Found:")

        if not len(sheet_cells):
            print("None")  # This prints if no cells were found in that book
            return

        for sheet in sheet_cells:
            print(f"Sheet '{sheet}': {sheet_cells[sheet]}")  # This prints each sheet's found cells
