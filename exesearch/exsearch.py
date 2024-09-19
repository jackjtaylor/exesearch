"""
This searches through worksheets by Excel to find search terms.

:return: The cells, sheets and books a term was found in
:rtype: print()
"""

import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from pathlib import Path
from tkinter import filedialog, Tk
from io import BytesIO
from msoffcrypto import OfficeFile
from collections import defaultdict
from warnings import filterwarnings


class Search:
    """
    This class holds properties related to a search.
    """

    def __init__(self) -> None:
        self.timestamp = datetime.datetime.now()
        self.found = defaultdict(lambda: "")
        self.found["test"] = "A1, A2"

    # This defines type-hints for later variables to be added.
    term: str
    path: Path
    is_exclusive: bool
    is_case_sensitive: bool
    found: defaultdict

    def get_found_count(self) -> int:
        """
        This function returns the count of how many matches were found.

        :return: How many matches were found.
        :rtype: int
        """
        count = 0
        for value in self.found.values():
            count += len(value.strip().split(","))
        return count


def create_search() -> Search:
    """
    This function requests a path from the user to search through. This then exclusively or
    inclusively searches for a term.

    :return: The search information, along with a timestamp.
    :rtype: Search
    """
    new_search = Search()

    new_search.path = get_search_directory()
    new_search.term = input("What term are you looking for? ")

    new_search.is_exclusive = get_exclusivity()
    new_search.is_case_sensitive = True

    if not new_search.is_exclusive:
        # This is asked only when an inclusive search is used
        new_search.is_case_sensitive = get_case_sensitivity()

    return new_search


def find_workbooks(search: Search):
    """
    This function searches through each file in a path and if valid, searches that workbook for the
    search term.

    :param search: The search to perform.
    :type search: Search
    """
    for file in search.path.rglob("*.xlsm" and "*.xlsx"):
        if file.is_file and "$" not in file.name:
            file_path = Path(search.path, file)
            search_for_term_in_workbook(file_path, search)


def get_case_sensitivity() -> bool:
    """
    This function asks the user if the search should be case sensitive.

    :return: If the search is case sensitive.
    :rtype: bool
    """
    return input("Is the term case sensitive? (y/n): ").lower().strip() == "y"


def get_exclusivity() -> bool:
    """
    This function asks the user if the search should be exclusive or not.

    :return: If the search should be exclusive
    :rtype: bool
    """
    return (
        input("Would you like to exclusively search, finding only exactly matching cells? (y/n): ")
        .lower()
        .strip()
        == "y"
    )  # This adds the count of terms found to the count, after searching a workbook.


def get_search_directory():
    """
    This function returns a searching directory chosen by the user.

    :return: The directory to search
    :rtype: Path
    """
    print("Please choose a directory.")

    root = Tk()  # This creates a hidden tkinter root.
    root.withdraw()

    path = Path(filedialog.askdirectory(title="Please choose a directory."))  # This shows a file
    # dialogue.

    return path


def prepare_workbook(book_path: Path) -> BytesIO:
    """
    This function prepares a workbook to be edited, by unlocking any encrypted files.

    :param book_path: The path of the workbook to prepare
    :type book_path: Path
    :return: The bytes stream of the decrypted workbook
    :rtype: BytesIO
    """
    with open(book_path, "rb") as workbook:
        office_file = OfficeFile(workbook)
        workbook.seek(0)  # This resets the reader for the next workbook.

        unencrypted_type = "plain"

        if office_file.is_encrypted and office_file.type != unencrypted_type:
            return decrypt_workbook(office_file)

        else:
            return BytesIO(workbook.read())  # If the workbook is unencrypted, create a stream.


def decrypt_workbook(office_file: OfficeFile) -> BytesIO:  # type: ignore
    """
    This is a function that decrypts a workbook using a password provided by the user. If the user
    doesn't enter the correct password after three attempts, the workbook is returned empty.

    :param office_file: The file to decrypt.
    :type office_file: OfficeFile
    :return: The decrypted file's byte stream.
    :rtype: OfficeFile
    """
    decrypted_workbook = BytesIO()  # This creates an in-memory BytesIO object to write the file to.

    for _ in range(3):  # This gives the user three tries to decrypt the workbook.
        key = input("Please enter the password to decrypt a file.")

        office_file.load_key(password=key)
        office_file.decrypt(decrypted_workbook)

        if decrypted_workbook:  # If the file was decrypted.
            break

    return decrypted_workbook  # This return the in-memory file decrypted.


def search_for_term_in_workbook(file_path: Path, search: Search):
    """
    This function finds terms inside a book, in all sheets.

    :param book: The path to the book
    :type book: Path
    :return: The count of how many times a term was found
    :rtype: int
    """
    workbook_byte_stream = prepare_workbook(file_path)

    workbook: Workbook = load_workbook(
        filename=workbook_byte_stream, data_only=True, read_only=True
    )
    print_workbook(file_path)  # This prints the name and path of the workbook

    for sheet in workbook:
        search_for_term_in_sheet(sheet, search)

    print_found_cells(search.found)


def search_for_term_in_sheet(sheet: Worksheet, search: Search):
    """
    This function searches for a term inside a sheet, within a workbook.

    :param count: _description_
    :type count: _type_
    :param found_cells: _description_
    :type found_cells: _type_
    :param sheet: _description_
    :type sheet: _type_
    """
    for column in sheet.iter_rows():
        for cell in column:
            if search.is_exclusive:  # If the search is exclusive, match the exact term.
                if search.term == str(cell.value):
                    add_to_found(sheet, cell, search)
            else:
                if search.is_case_sensitive:  # If the search is just case sensitive, match words.
                    if search.term in str(cell.value):
                        add_to_found(sheet, cell, search)
                else:  # If the search only needs the same characters, match anything.
                    if search.term.lower().strip() in str(cell.value).lower().strip():
                        add_to_found(sheet, cell, search)


def add_to_found(sheet: Worksheet, cell: Cell, search: Search):
    """
    This function adds a cell to the found cells in a search.

    :param sheet: The worksheet the cell was found in.
    :type sheet: Worksheet
    :param cell: The cell that was found to match.
    :type cell: Cell
    :param search: The search to add this result to.
    :type search: Search
    """
    search.found[sheet.title] += f"{cell.column_letter}{cell.row}, "


def print_workbook(book_path: Path) -> None:
    """
    This function prints the workbook name and path.

    :param book_path: The path to the workbook
    :type book_path: Path
    """
    print(f"\n_______________ Book: {book_path.name} _______________\n")
    print(f"Path: {book_path.absolute()}\n")


def print_found_cells(sheet_cells: defaultdict[str, str]) -> None:
    """
    This function prints the cells found, grouped by sheet.

    :param sheet_cells: The cells found, by sheet
    :type sheet_cells: defaultdict[str, str]
    """
    print("Cells Found:")

    if not len(sheet_cells):
        print("None")  # This prints if no cells were found in that book
        return

    for sheet in sheet_cells:
        print(f"Sheet '{sheet}': {sheet_cells[sheet]}")  # This prints each sheet's found cells


def main():
    filterwarnings("ignore", category=UserWarning, module="openpyxl")

    search = create_search()
    find_workbooks(search)

    input("\nPress enter to exit.")


if __name__ == "__main__":
    main()
